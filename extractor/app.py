"""
Extractor sidecar — two endpoints used by n8n:

POST /extract-name   { "pdf_path": "/data/files/assignments/foo.pdf" }
  → { "name": "John Smith" }

POST /match-roster   { "name": "John Smith", "roster_path": "/data/files/roster.xlsx",
                       "name_col": "Name", "email_col": "Email" }
  → { "matched_name": "John Smith", "email": "john@uni.edu" }

POST /find-review    { "name": "John Smith", "reviews_dir": "/data/files/reviews" }
  → { "path": "/data/files/reviews/John Smith review.docx" }  or  { "path": null }
"""

import os
import json
from pathlib import Path

import pymupdf
import openpyxl
from flask import Flask, request, jsonify, send_file
from openai import OpenAI
from rapidfuzz import process, fuzz

app = Flask(__name__)
_template_cache: str | None = None
_roster_cache: dict[str, dict] = {}  # path -> roster dict
# Configurable backend: set LLM_BASE_URL + LLM_API_KEY in .env for any OpenAI-compatible endpoint
# e.g. Gemini: LLM_BASE_URL=https://generativelanguage.googleapis.com/v1beta/openai/
#      local:  LLM_BASE_URL=http://host.docker.internal:11434/v1  LLM_API_KEY=ollama
_llm_base = os.environ.get("LLM_BASE_URL")
_llm_key  = os.environ.get("LLM_API_KEY") or os.environ.get("OPENAI_API_KEY", "")
client = OpenAI(base_url=_llm_base, api_key=_llm_key) if _llm_base else OpenAI(api_key=_llm_key)


# ── helpers ────────────────────────────────────────────────────────────────────

def file_to_text(file_path: str) -> str:
    """Extract text from first 2 pages of PDF or DOCX (~1500 chars max)."""
    path = Path(file_path)
    if path.suffix.lower() == ".docx":
        from docx import Document
        from docx.oxml.ns import qn
        doc = Document(str(path))
        lines, page = [], 1
        for p in doc.paragraphs:
            # detect page break
            for br in p._element.iter(qn('w:lastRenderedPageBreak')):
                page += 1
            for br in p._element.iter(qn('w:pageBreak')):
                page += 1
            if page > 2:
                break
            if p.text.strip():
                lines.append(p.text.strip())
        text = "\n".join(lines)
    else:
        doc = pymupdf.open(str(path))
        pages = min(2, len(doc))
        text = "\n".join(doc[i].get_text("text").strip() for i in range(pages))
        doc.close()
    return text[:1500]


def extract_name_from_filename(file_path: str) -> str:
    """Moodle filenames: 'First Last_studentid_assignsubmission_...' — name is before first underscore-digit."""
    stem = Path(file_path).stem
    import re
    m = re.match(r'^(.+?)_\d', stem)
    return m.group(1).strip() if m else stem.split('_')[0].strip()


def page1_to_base64(file_path: str) -> str | None:
    """Render first page of PDF or DOCX to a base64 PNG for vision fallback."""
    import base64, tempfile
    path = Path(file_path)
    try:
        if path.suffix.lower() == ".docx":
            # pymupdf can open docx directly since 1.23
            doc = pymupdf.open(str(path))
        else:
            doc = pymupdf.open(str(path))
        if not doc.page_count:
            return None
        pix = doc[0].get_pixmap(dpi=150)
        doc.close()
        return base64.b64encode(pix.tobytes("png")).decode()
    except Exception as e:
        app.logger.warning(f"page1_to_base64 failed: {e}")
        return None


def gpt_extract_name(text: str, image_b64: str | None = None) -> dict:
    system_msg = (
        "You are a document parser. Extract student names and paper title from the content. "
        "Return ONLY valid JSON: {\"names\": [...], \"paper_title\": \"...\"}. No markdown fences."
    )
    if image_b64:
        user_content = [
            {"type": "text", "text": "Extract student full names and paper title from this cover page image."},
            {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{image_b64}"}},
        ]
        if text:
            user_content[0]["text"] += f"\n\nAlso available as text:\n{text[:500]}"
    else:
        user_content = f"Document text:\n\n{text[:2000]}"

    resp = client.chat.completions.create(
        model=os.environ.get("LLM_MODEL", "gpt-4o-mini"),
        messages=[
            {"role": "system", "content": system_msg},
            {"role": "user", "content": user_content},
        ],
        max_tokens=150,
        temperature=0,
    )
    content = resp.choices[0].message.content
    if content is None:
        parts = getattr(resp.choices[0].message, 'parts', None)
        content = parts[0].text if parts else "{}"
    content = content.strip().strip("```json").strip("```").strip()
    try:
        return json.loads(content)
    except Exception:
        return {"names": [n.strip() for n in content.split(",") if n.strip()],
                "paper_title": ""}


def load_roster(xlsx_path: str, name_col: str = "", email_col: str = "") -> dict[str, dict]:
    global _roster_cache
    if xlsx_path in _roster_cache:
        return _roster_cache[xlsx_path]
    import re
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(max_row=1))]
    lower_headers = [x.lower() for x in headers]

    # Auto-detect paired "Student Name N" / "Student Email N" columns
    pairs = []
    for i, h in enumerate(headers):
        m = re.match(r"Student Name (\d+)$", h, re.IGNORECASE)
        if m:
            email_header = f"Student Email {m.group(1)}"
            if email_header.lower() in lower_headers:
                pairs.append((i, lower_headers.index(email_header.lower())))

    # Fallback to explicit col names
    if not pairs and name_col and email_col:
        pairs = [(lower_headers.index(name_col.lower()), lower_headers.index(email_col.lower()))]

    if not pairs:
        raise ValueError(f"No Student Name/Email column pairs found. Headers: {headers}")

    # Supervisor columns (optional)
    sup_name_col  = lower_headers.index("supervisor")       if "supervisor"       in lower_headers else None
    sup_email_col = lower_headers.index("supervisor email") if "supervisor email" in lower_headers else None

    roster = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        sup_name  = str(row[sup_name_col]).strip()  if sup_name_col  is not None and row[sup_name_col]  else ""
        sup_email = str(row[sup_email_col]).strip() if sup_email_col is not None and row[sup_email_col] else ""
        for ni, ei in pairs:
            name  = str(row[ni]).strip() if row[ni] else ""
            email = str(row[ei]).strip() if row[ei] else ""
            if name and email and name.lower() != "none":
                clean_sup = re.sub(r'^(prof\.?\s+|dr\.?\s+)+', '', sup_name, flags=re.IGNORECASE).strip()
                roster[name.title()] = {"email": email, "supervisor_name": clean_sup.title() if clean_sup else "", "supervisor_email": sup_email}
    _roster_cache[xlsx_path] = roster
    return roster


def fuzzy_best(query: str, candidates: list[str], threshold: int = 65, scorer=None) -> str | None:
    scorer = scorer or fuzz.token_sort_ratio
    q = query.lower()
    lower_map = {c.lower(): c for c in candidates}
    result = process.extractOne(q, list(lower_map.keys()), scorer=scorer)
    if result and result[1] >= threshold:
        return lower_map[result[0]]
    return None


# ── routes ─────────────────────────────────────────────────────────────────────

@app.post("/clear-cache")
def clear_cache():
    global _roster_cache, _template_cache
    _roster_cache = {}
    _template_cache = None
    return jsonify({"ok": True})


@app.get("/list-assignments")
def list_assignments():
    assignments_dir = Path("/data/files/assignments")
    files = [
        str(f) for f in assignments_dir.iterdir()
        if f.suffix.lower() in (".pdf", ".docx")
    ] if assignments_dir.exists() else []
    return jsonify({"files": sorted(files)})



@app.post("/extract-name")
def extract_name():
    file_path = request.json.get("pdf_path") or request.json.get("file_path")
    if not file_path or not Path(file_path).exists():
        return jsonify({"error": f"File not found: {file_path}"}), 400
    text = file_to_text(file_path)
    # ponytail: 100-char threshold; if cover page is an image, text will be near-empty
    image_b64 = page1_to_base64(file_path) if len(text.strip()) < 100 else None
    if image_b64:
        app.logger.info(f"extract-name: text too short ({len(text)} chars), using vision fallback")
    result = gpt_extract_name(text, image_b64=image_b64)
    names = result.get("names") or []
    if isinstance(names, str):
        names = [n.strip() for n in names.split(",") if n.strip()]
    names = [n.title() for n in names if n]
    # Vision fallback: if LLM returned no names and we haven't tried image yet, retry with page image
    if not names and image_b64 is None:
        app.logger.info("extract-name: no names from text, retrying with vision fallback")
        image_b64 = page1_to_base64(file_path)
        if image_b64:
            result = gpt_extract_name(text, image_b64=image_b64)
            names = result.get("names") or []
            if isinstance(names, str):
                names = [n.strip() for n in names.split(",") if n.strip()]
            names = [n.title() for n in names if n]
    return jsonify({
        "name": names[0] if names else "",
        "names": names,
        "paper_title": result.get("paper_title", ""),
        "page1_text": text[:500],
    })


@app.post("/match-roster")
def match_roster():
    body = request.json
    # Accept either a single name or a list of names; try each until a match is found
    raw_names  = body.get("names") or [body["name"]]
    # n8n may send the array as a JSON string — parse it
    if isinstance(raw_names, str):
        import ast
        try:
            raw_names = json.loads(raw_names)
        except Exception:
            try:
                raw_names = ast.literal_eval(raw_names)
            except Exception:
                raw_names = [n.strip() for n in raw_names.split(",") if n.strip()]
    names = [n for n in raw_names if n]
    roster_path= body["roster_path"]
    name_col   = body.get("name_col", "")
    email_col  = body.get("email_col", "")
    threshold  = int(body.get("threshold", 65))

    roster = load_roster(roster_path, name_col, email_col)
    matched_names, supervisor_emails, supervisor_names = [], [], []
    sup_email_seen = set()
    for name in names:
        matched = fuzzy_best(name, list(roster.keys()), threshold)
        if matched:
            matched_names.append(matched)
            entry = roster[matched]
            sup_email = entry.get("supervisor_email", "")
            sup_name  = entry.get("supervisor_name", "")
            if sup_email and sup_email not in sup_email_seen:
                supervisor_emails.append(sup_email)
                sup_email_seen.add(sup_email)
            if sup_name and sup_name not in supervisor_names:
                supervisor_names.append(sup_name)
    if not matched_names:
        return jsonify({"error": f"No roster match for {names}"}), 404
    return jsonify({
        "matched_name":    matched_names[0],
        "matched_names":   matched_names,
        "to":              ", ".join(supervisor_emails),
        "supervisor_name": ", ".join(supervisor_names),
    })


@app.post("/find-review")
def find_review():
    name        = request.json["name"]
    reviews_dir = request.json["reviews_dir"]

    docx_files = list(Path(reviews_dir).glob("*.docx"))
    if not docx_files:
        return jsonify({"path": None})

    stems = {f.stem: f for f in docx_files}
    # Fuzzy match student name against file stems — no LLM needed
    # partial_ratio handles Moodle's garbled filenames (name appears as substring)
    matched_stem = fuzzy_best(name, list(stems.keys()), threshold=70, scorer=fuzz.partial_ratio)
    path = stems.get(matched_stem) if matched_stem else None
    return jsonify({"path": str(path) if path else None})


@app.get("/download-review")
@app.get("/download-assignment")
def download_assignment():
    """Serve an assignment file for n8n to attach to Gmail draft."""
    file_path = request.args.get("path")
    app.logger.info(f"download-assignment: path={file_path!r}")
    if not file_path or not Path(file_path).exists():
        app.logger.error(f"download-assignment: NOT FOUND path={file_path!r}")
        return jsonify({"error": f"not found: {file_path}"}), 404
    return send_file(file_path, as_attachment=True, download_name=Path(file_path).name)


@app.get("/get-template")
def get_template():
    global _template_cache
    if _template_cache is None:
        # Prefer .md; fall back to .txt
        md_path  = Path("/data/files/email_template.md")
        txt_path = Path("/data/files/email_template.txt")
        p = md_path if md_path.exists() else txt_path
        raw = p.read_text() if p.exists() else ""
        _template_cache = raw
    return jsonify({"template": _template_cache})


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5050)
